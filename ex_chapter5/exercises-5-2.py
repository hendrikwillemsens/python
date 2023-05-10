#!/usr/bin/python3
################################################################################
# 5.2 Get sample IDs and file sizes from folder with htseqcount files
################################################################################
# Import all required modules
import os
import re
import humanfriendly
from openpyxl import Workbook

################################################################################
# V1
# Initialize Workbook
wb = Workbook()
# Active worksheet
ws = wb.active
# Add header row
ws.append(["sample_id", "filesize"])
# Scan folder

# Save workbook file
wb.save("sample_ids_sizes_V1.xlsx")



################################################################################
# V2
# Initialize Workbook
wb = Workbook()
# Active worksheet
ws = wb.active
# Add header row
ws.append(["sample_id", "filesize"])
# Scan folder: exercise-5-2-htseqcount
location = os.getcwd()
ex52dir = location + "/exercise-5-2-htseqcount/"
file_list = os.listdir(ex52dir)
print("Files and folders in {} :".format(ex52dir))
for file in file_list:
    print("Current file: ", file)
    ws.append([file])
# Save workbook file
wb.save("sample_ids_sizes_V2.xlsx")



################################################################################
# V3

# Initialize Workbook
wb = Workbook()
# Active worksheet
ws = wb.active
# Add header row
ws.append(["sample_id", "filesize"])
# Scan folder: exercise-5-2-htseqcount
location = os.getcwd()
ex52dir = location + "/exercise-5-2-htseqcount/"
file_list = os.listdir(ex52dir)
print("Files and folders in {} :".format(ex52dir))
for file in file_list:
    # print("Current file: ", file)
    result = re.search("^count", file)
    if result:
        # print("is a count file!")
        sample_list = file.split('_')
        sample_id = sample_list[1]
        print("\nsample_id = ", sample_id)
        # Concatinate directory and filename
        size_bytes = os.path.getsize(ex52dir + file)
        hsize = humanfriendly.format_size(size_bytes, binary=True)
        print("filesize = {} ({})".format(hsize,size_bytes))
        # Add row with sample_id and filesize in worksheet
        ws.append([sample_id, hsize])
# Save workbook file
wb.save("sample_ids_sizes.xlsx")
