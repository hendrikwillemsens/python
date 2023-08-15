#!/usr/bin/python3

import re
import os
import humanfriendly
from openpyxl import Workbook
wb = Workbook()
ws = wb.active

location = "/media/sf_SF/python/ex_chapter5/"
os.chdir(location)
print(os.getcwd())

htseqcount = location + "exercise-5-2-htseqcount/"
htseq_list = os.listdir(htseqcount)


ws["A1"] = "sample_id"
ws["B1"] = "filesize"

count = 2
for file in htseq_list:  
    result = re.search("filtered", file)
    if result:
        lijst = file.split("_")
        print("sample_id = {}" . format(lijst[1]))
        colom= "A" + str(count)
        ws[colom] = lijst[1]
        

        rij = "B" + str(count)
        
        size = htseqcount + file      
        size_bytes = os.path.getsize(size)
        hsize = humanfriendly.format_size(size_bytes, binary=True)
        print("file size = {}".format(hsize))
        ws[rij] = str(hsize)
        print("\t")
        count = count + 1
        
wb.save('oefening5.2.xlsx')    
           

 