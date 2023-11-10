#!/usr/bin/python3

from openpyxl import Workbook
from openpyxl.styles import Side, Border, Alignment
import csv
import re
from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws['A1'] = 'HMM'
ws['B1'] = 'target'
ws['C1'] = 'E-value'
ws['D1'] = 'score'
ws['E1'] = '#domains'
ws['F1'] = 'Sequence'
ws['G1'] = 'Description'



fileObj = open("LysM.hmm-GCA_000025205.1_ASM2520v1.txt", "r")
c = 1
print("Reading file LysM.hmm-GCA_000025205.1_ASM2520v1.txt")
for line in fileObj.readlines():
    result = re.search("^#", line)
    
    if result:
        resultaat = re.search('faa$', line)
        query = re.search('.hmm$', line)
        if resultaat:
            array = line.split(" ")
            print("Target sequence = {}" .format(array[11]))
            
        if query:
            file = line.split(" ")
            print("HMM file = {}".format(file[21]))        
fileObj.close()


txt2 = open("LysM.hmm-GCA_000159155.2_ASM15915v2.txt", "r")
c = 1
print("Reading file LysM.hmm-GCA_000159155.2_ASM15915v2")
for line in txt2.readlines():
    result = re.search("^#", line) 
    if result:
        resultaat = re.search('faa$', line)
        query = re.search('.hmm$', line)
        
        if resultaat:
            array = line.split(" ")
            print("Target sequence = {}" .format(array[11]))
            
        if query:
            file = line.split(" ")
            print("HMM file = {}".format(file[21]))      
txt2.close()

txt3 = open("LysM.hmm-GCA_002871555.1_ASM287155v1.txt", "r")
c = 1
print("Reading file LysM.hmm-GCA_002871555.1_ASM287155v1")
for line in txt3.readlines():
    result = re.search("^#", line)
    
    if result:
        resultaat = re.search('faa$', line)
        query = re.search('.hmm$', line)
        if resultaat:
            array = line.split(" ")
            print("Target sequence = {}" .format(array[11]))
            
        if query:
            file = line.split(" ")
            print("HMM file = {}".format(file[21]))      
txt3.close()

txt4 = open("Thiol_cytolys_C.hmm-GCA_004336715.1_ASM433671v1.txt", "r")
c = 1
print("Reading file Thiol_cytolys_C.hmm-GCA_004336715.1_ASM433671v1.txt")
for line in txt4.readlines():
    result = re.search("^#", line)
    
    if result:
        resultaat = re.search('faa$', line)
        query = re.search('.hmm$', line)
        if resultaat:
            array = line.split(" ")
            print("Target sequence = {}" .format(array[11]))
            
        if query:
            file = line.split(" ")
            print("HMM file = {}".format(file[21]))      
txt4.close()

txt5 = open("Thiol_cytolysin.hmm-GCA_000025205.1_ASM2520v1.txt", "r")
c = 1
print("Reading file Thiol_cytolysin.hmm-GCA_000025205.1_ASM2520v1")
for line in txt5.readlines():
    result = re.search("^#", line)
    
    if result:
        resultaat = re.search('faa$', line)
        query = re.search('.hmm$', line)
        if resultaat:
            array = line.split(" ")
            print("Target sequence = {}" .format(array[11])) 
        if query:
            file = line.split(" ")
            print("HMM file = {}".format(file[21]))      
txt5.close()

wb.save('hmm-excel.xlsx')