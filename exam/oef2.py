#!/usr/bin/python3
import csv
from openpyxl import Workbook
from openpyxl.styles import Side, Border, Alignment
wb = Workbook()
ws = wb.active




mydict = {}
lijst1=[]
lijst2=[]
lijst3 = []
c = 1
x = 1
"""
with open('NCBI_Genome_result_Legionella.txt') as filedata:
    lines = filedata.readlines()
    for line in lines:
        if line.startswith(str(c)):
            hendrik = line.split(".")
            wijn = hendrik[1].strip("\n")                            
            lijst1.append(wijn)          
            print ("{}: {}".format(c, wijn))
            ws['c{}'.format(counter)]= "ncbi.gov/genome/?term=" + wijn
            c = c + 1
            
        if line.startswith("Genome"):
            bier = line.split()
          
            lijst2.append(bier[2])
    fruit = dict(zip(lijst1, lijst2))  
    
    
    print("Data from NCBI Genome file stored:\n ", fruit)
    """

with open('lpsn_export.csv') as csv_file:
    csv_reader = csv.reader(csv_file, delimiter=',')
    line_count = 1
    counter = 2
    for row in csv_reader:
        if row[0] == "Legionella":
            print("Found: ({}) {} {} \n -->{}".format(line_count, row[0], row[1], row[6]))
            line_count= line_count + 1
            ws.column_dimensions["A"].width = 50.0
            ws.column_dimensions["B"].width = 50.0
            ws.column_dimensions["C"].width = 50.0
            ws['A{}'.format(counter)] = row[0] + " " + row[1]
            ws['B{}'.format(counter)] = row[6]
            
            
            counter = counter + 1       
d = 2

with open('NCBI_Genome_result_Legionella.txt') as filedata:
    lines = filedata.readlines()
    for line in lines:
        if line.startswith(str(c)):
            hendrik = line.split(".")
            wijn = hendrik[1].strip("\n")                            
            lijst1.append(wijn)          
            print ("{}: {}".format(c, wijn))
            
            c = c + 1
            
        if line.startswith("Genome"):
            bier = line.split()
            ws['c{}'.format(d)]= "ncbi.gov/genome/?term=" + bier[2]
            d = d + 1
            lijst2.append(bier[2])
    fruit = dict(zip(lijst1, lijst2))  











wb.save('oef2.xlsx')