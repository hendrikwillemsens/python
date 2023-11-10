import MySQLdb as my
from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws['A1'] = "Accession no."
ws['B1'] = "SNP"
ws['C1'] = "Chr:start-stop"

accession = []
while True:
    response = input("Enter accession no. or type 0 to stop: ")   
    if response != str(0):
        accession.append(response)
    else: 
        print(accession)
        break

db = my.connect(host="genome-mysql.soe.ucsc.edu",
   user="genomep",
   passwd="password",
   db="hg38")
c = db.cursor()       

counter = 2
for i in accession:
    acesie = str("SELECT * FROM snp151CodingDbSnp where transcript = '{}'".format(i))
    print("using query: {}" .format(acesie))
    no_rows = c.execute("""{}""".format(acesie))
    result = c.fetchall()
    print("Total rows ensGene: {}".format(no_rows))
    for row in result:        
        ws.append([row[0], row[1]])    
        ws['A{}'.format(counter)]= i
        ws['B{}'.format(counter)]= row[1]
        ws['C{}'.format(counter)]= row[2]
        counter = counter + 1
    
    

wb.save('voorbeelexamenoef1.xlsx')